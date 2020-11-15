import json
import re
import openpyxl
from openpyxl import load_workbook, Workbook

class XlCell(object):
    merged = False
    merged_ignore = False
    merged_range = None
    font = None
    border = None
    fill = None
    number_format = None
    protection = None
    alignment = None
    value = None
    col_idx = None
    has_style = False
    is_date = None
    data_type = None
    encoding = None
    
    def assign_from(self,cell):
        for attr in dir(self):
            if not attr.startswith('__') and not callable(getattr(self,attr)):
                if attr not in ['merged','merged_ignore', 'merged_range']:
                    if hasattr(cell,attr):
                        setattr(self,attr,getattr(cell,attr))

        
def copy_sheet_properties(fromsheet, tosheet):
    attr_dict = [
                    'column_dimensions','page_margins','page_setup'  
                ]
    for attr in attr_dict:
        setattr(tosheet,attr,getattr(fromsheet,attr))
        
class XlRender(object):
    read_book = None
    read_sheet = None
    mapping_properties = None
    write_book = None
    write_sheet = None
    
    def __init__(self, template, deffile ):
        self.mapping_properties = json.load(open(deffile))
        self.read_book = load_workbook(template)
        self.read_sheet = self.read_book.active
        
        self.write_book = Workbook()
        self.write_sheet = self.write_book.active
        copy_sheet_properties(self.read_sheet, self.write_sheet)
    
    def _get_area(self,name, data=None):
        rows = []
        cellrange = None
        parameters = []
        if data is not None:
            data = json.loads(data)
        
        for area in self.mapping_properties:
            if area["area_name"] == name:
                start_row = area["start_row"]
                end_row = area["end_row"]
                last_column = self.read_sheet.max_column
                # cellrange = self.read_sheet.get_squared_range(1,start_row,last_column,end_row)

                cellrange = self.read_sheet.iter_rows(min_col=1, min_row=start_row, max_col=last_column, max_row=end_row)


                if 'parameters' in area:
                    parameters = area["parameters"]
                break
            
        #    get row
        if cellrange is not None:
            for row in cellrange:
                xrow = []
                for cell in row:
                    xlcell = XlCell()
                    xlcell.assign_from(cell)
                    
                    #print xlcell.row
                    idx = cell.coordinate
                    
                    for merged_range in self.read_sheet.merged_cell_ranges:
                        print("merged_range", merged_range)
                        merged_cells = list(openpyxl.utils.rows_from_range(str(merged_range)))
                        processed = False
                        #print merged_cells , "merged_cell", merged_cells[0][0]
                        for mgrow in merged_cells:
                            if idx in mgrow:
                                processed = True
                                xlcell.merged = True
                                if idx != merged_cells[0][0] :
                                    xlcell.merged_ignore = True
                                else:
                                    first_cell = merged_cells[0][0]
                                    first_cell_split = re.findall('\d+|\D+', first_cell)
                                    first_cell_col_idx = first_cell_split[0]
                                    first_cell_row = first_cell_split[1]
                                   
                                    last_row = merged_cells[len(merged_cells) -1]
                                    last_cell = last_row[len(last_row) - 1]
                                    last_cell_split = re.findall('\d+|\D+', last_cell)
                                    last_cell_col_idx = last_cell_split[0]
                                    last_cell_row = last_cell_split[1]
                                    
                                    merged_range = {
                                                        "start_col": first_cell_col_idx,
                                                        "end_cold":last_cell_col_idx,
                                                        "row_span": int(last_cell_row) - int(first_cell_row)
                                                    }
                                    xlcell.merged_range = merged_range
                                    
                                break
                        if processed:
                            break
                    
                    if (not xlcell.merged_ignore) and (data is not None) and (len(parameters) > 0):
                        for param in parameters:
                            if param['position'] == idx:
                                if ( param['name'] in data):
                                    xlcell.value = data[param['name']]

                    xrow.append(xlcell)
                rows.append(xrow)            
            return rows        
        return None
    
    def put_area(self, name, data=None):

        cellrange = self._get_area(name,data=data)
        
        if (cellrange is None) or (self.write_book is None) or (self.write_sheet is None):
            return
        
        max_row = self.write_sheet.max_row
        max_col = self.write_sheet.max_column
        # frist_range = self.write_sheet.get_squared_range(1,max_row,max_col,max_row)

        frist_range = self.read_sheet.iter_rows(min_col=1, min_row=max_row, max_col=max_col, max_row=max_row)

        
        last_row_has_data = False
        for row in frist_range:
            for cell in row:
                if (cell.value is not None) or (cell.has_style):
                    last_row_has_data = True
                    break
            if last_row_has_data:
                break
        if last_row_has_data:
            max_row = max_row + 1
        
        
        for xrow in cellrange:
            for xlcell in xrow:
                new_cell = self.write_sheet.cell(row=max_row, column=xlcell.col_idx)
                if not xlcell.merged_ignore :
                    if (xlcell.merged):
                        mgrange = xlcell.merged_range['start_col'] + str(max_row) + ":" \
                                + xlcell.merged_range['end_cold'] + str(max_row + xlcell.merged_range['row_span'])
                        self.write_sheet.merge_cells(mgrange)
                    new_cell.value = xlcell.value
                # if xlcell.has_style:
                #     new_cell.font = xlcell.font
                #     new_cell.border = xlcell.border
                #     new_cell.fill = xlcell.fill
                #     new_cell.number_format = xlcell.number_format
                #     new_cell.protection = xlcell.protection
                #     new_cell.alignment = xlcell.alignment
            max_row = max_row + 1
            
    def save(self, out):
        if self.write_book is not None:
            self.write_book.save(out)